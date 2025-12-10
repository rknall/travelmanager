# SPDX-FileCopyrightText: 2025 Roland Knall <rknall@gmail.com>
# SPDX-License-Identifier: GPL-2.0-only
"""Expense report generator service."""

import io
import zipfile
from datetime import datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from slugify import slugify
from sqlalchemy.orm import Session

from src.integrations.base import DocumentProvider
from src.models import Event, Expense
from src.services import expense_service, integration_service


def _slugify_filename(name: str, max_length: int = 50) -> str:
    """Create a slug suitable for filenames."""
    slug = slugify(name, lowercase=True, separator="_")
    return slug[:max_length]


def _format_date(d: Any) -> str:
    """Format a date for filenames."""
    if hasattr(d, "strftime"):
        return d.strftime("%Y-%m-%d")
    return str(d)


class ExpenseReportGenerator:
    """Generator for expense reports with Excel and document packaging."""

    def __init__(
        self,
        db: Session,
        paperless: DocumentProvider | None = None,
    ) -> None:
        """Initialize the expense report generator.

        Args:
            db: Database session for querying expenses.
            paperless: Optional Paperless provider for downloading documents.
        """
        self.db = db
        self.paperless = paperless

    def get_preview(self, event: Event) -> dict[str, Any]:
        """Return summary without generating files."""
        expenses = expense_service.get_expenses(self.db, event.id)

        total = sum(e.amount for e in expenses)
        documents_available = sum(1 for e in expenses if e.paperless_doc_id)

        by_category: dict[str, Decimal] = {}
        by_payment_type: dict[str, Decimal] = {}

        for expense in expenses:
            cat = expense.category.value
            by_category[cat] = by_category.get(cat, Decimal(0)) + expense.amount

            pt = expense.payment_type.value
            by_payment_type[pt] = by_payment_type.get(pt, Decimal(0)) + expense.amount

        return {
            "event_id": event.id,
            "event_name": event.name,
            "company_name": event.company.name if event.company else None,
            "start_date": _format_date(event.start_date),
            "end_date": _format_date(event.end_date),
            "expense_count": len(expenses),
            "documents_available": documents_available,
            "total": float(total),
            "currency": expenses[0].currency if expenses else "EUR",
            "by_category": {k: float(v) for k, v in by_category.items()},
            "by_payment_type": {k: float(v) for k, v in by_payment_type.items()},
            "paperless_configured": self.paperless is not None,
        }

    def _create_excel(
        self,
        event: Event,
        expenses: list[Expense],
    ) -> bytes:
        """Create Excel spreadsheet for expenses."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        currency_format = '#,##0.00 "€"'
        date_format = "YYYY-MM-DD"

        # Title row
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value = f"Expense Report: {event.name}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        # Info rows
        ws["A2"] = f"Company: {event.company.name if event.company else 'N/A'}"
        start = _format_date(event.start_date)
        end = _format_date(event.end_date)
        ws["A3"] = f"Period: {start} to {end}"
        ws["A4"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Headers
        headers = [
            "#",
            "Date",
            "Description",
            "Category",
            "Payment Type",
            "Amount",
            "Document",
        ]
        header_row = 6
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Data rows
        total = Decimal(0)
        for idx, expense in enumerate(expenses, 1):
            row = header_row + idx
            ws.cell(row=row, column=1, value=idx).border = border
            date_cell = ws.cell(row=row, column=2, value=expense.date)
            date_cell.number_format = date_format
            date_cell.border = border
            ws.cell(row=row, column=3, value=expense.description or "").border = border
            ws.cell(row=row, column=4, value=expense.category.value).border = border
            ws.cell(row=row, column=5, value=expense.payment_type.value).border = border
            amount_cell = ws.cell(row=row, column=6, value=float(expense.amount))
            amount_cell.number_format = currency_format
            amount_cell.border = border
            doc_ref = f"{idx:02d}_*.pdf" if expense.paperless_doc_id else "N/A"
            ws.cell(row=row, column=7, value=doc_ref).border = border
            total += expense.amount

        # Total row
        total_row = header_row + len(expenses) + 1
        ws.cell(row=total_row, column=5, value="Total:").font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=6, value=float(total))
        total_cell.font = Font(bold=True)
        total_cell.number_format = currency_format

        # Adjust column widths
        column_widths = [5, 12, 40, 15, 15, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def generate(self, event: Event) -> bytes:
        """Generate ZIP with Excel and documents."""
        expenses = expense_service.get_expenses(self.db, event.id)
        expenses.sort(key=lambda e: e.date)

        # Create the Excel file
        excel_bytes = self._create_excel(event, expenses)

        # Create ZIP file
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            # Add Excel file
            event_slug = _slugify_filename(event.name)
            date_str = datetime.now().strftime("%Y-%m-%d")
            excel_name = f"expense_report_{event_slug}_{date_str}.xlsx"
            zip_file.writestr(excel_name, excel_bytes)

            # Add documents from Paperless if available
            if self.paperless:
                for idx, expense in enumerate(expenses, 1):
                    if expense.paperless_doc_id:
                        try:
                            (
                                content,
                                original_name,
                                _mime_type,
                            ) = await self.paperless.download_document(
                                expense.paperless_doc_id
                            )
                            # Extract extension from original filename or mime type
                            ext = "pdf"
                            if "." in original_name:
                                ext = original_name.rsplit(".", 1)[-1].lower()

                            # Create standardized filename
                            desc_slug = _slugify_filename(
                                expense.description or "document", 30
                            )
                            date_fmt = _format_date(expense.date)
                            new_filename = f"{idx:02d}_{date_fmt}_{desc_slug}.{ext}"

                            zip_file.writestr(f"documents/{new_filename}", content)
                        except Exception:  # noqa: S110
                            # Skip documents that fail to download
                            pass

        zip_buffer.seek(0)
        return zip_buffer.getvalue()

    def get_filename(self, event: Event) -> str:
        """Get the filename for the ZIP file."""
        event_slug = _slugify_filename(event.name)
        date_str = datetime.now().strftime("%Y-%m-%d")
        return f"expense_report_{event_slug}_{date_str}.zip"


async def create_report_generator(
    db: Session,
    event: Event,
) -> ExpenseReportGenerator:
    """Create a report generator with optional Paperless provider."""
    paperless_config = integration_service.get_active_document_provider(db)
    paperless = None

    if paperless_config:
        paperless = integration_service.create_provider_instance(paperless_config)
        if not isinstance(paperless, DocumentProvider):
            paperless = None

    return ExpenseReportGenerator(db, paperless)
